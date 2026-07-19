"""Calculate recognition-memory metrics from E-Prime Excel exports.

Expected input layout::

    input_directory/
        sub01/
            off.xlsx
            10Hz.xlsx
            130Hz.xlsx
        sub02/
            ...

The script identifies retrieval trials, removes responded trials with reaction
times below 200 ms, classifies old/new responses, computes signal-detection
metrics and reaction-time summaries, and writes an Excel workbook containing
session-level, trial-level, exclusion, and quality-control records.
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path
from statistics import NormalDist
from typing import Any

import openpyxl
import pandas as pd

MIN_RT_MS = 200

CONDITION_FILES = {
    "off": "off.xlsx",
    "10Hz": "10Hz.xlsx",
    "130Hz": "130Hz.xlsx",
}
CONDITION_ORDER = ["off", "10Hz", "130Hz"]

REQUIRED_COLUMNS = [
    "Searching.ACC",
    "SearchingPicCRBox",
    "Searching.RESP",
    "Searching.RT",
]


def subject_sort_key(path: Path) -> tuple[float, str]:
    match = re.search(r"(\d+)", path.name)
    return (int(match.group(1)) if match else float("inf"), path.name)


def safe_divide(numerator: int, denominator: int) -> float:
    return numerator / denominator if denominator else float("nan")


def corrected_rate(count: int, total: int) -> float:
    if total == 0:
        return float("nan")

    rate = count / total
    if rate in (0, 1):
        return (count + 0.5) / (total + 1)
    return rate


def z_score_from_rate(rate: float) -> float:
    if pd.isna(rate):
        return float("nan")
    return NormalDist().inv_cdf(rate)


def summarize_reaction_time(series: pd.Series, prefix: str) -> dict[str, Any]:
    values = pd.to_numeric(series, errors="coerce").dropna()
    if values.empty:
        return {
            f"{prefix}_n": 0,
            f"{prefix}_mean": float("nan"),
            f"{prefix}_median": float("nan"),
            f"{prefix}_sd": float("nan"),
            f"{prefix}_min": float("nan"),
            f"{prefix}_max": float("nan"),
        }

    return {
        f"{prefix}_n": int(len(values)),
        f"{prefix}_mean": float(values.mean()),
        f"{prefix}_median": float(values.median()),
        f"{prefix}_sd": float(values.std(ddof=1)) if len(values) > 1 else float("nan"),
        f"{prefix}_min": float(values.min()),
        f"{prefix}_max": float(values.max()),
    }


def classify_trial(stimulus_type: float, response: float) -> str:
    if stimulus_type == 3:
        if response == 3:
            return "hit"
        if response == 4:
            return "miss"
        return "no_response"

    if stimulus_type == 4:
        if response == 3:
            return "false_alarm"
        if response == 4:
            return "correct_rejection"
        return "no_response"

    return "not_retrieval_trial"


def find_header_row(
    file_path: Path,
    required_columns: list[str],
    max_scan_rows: int = 20,
) -> int:
    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    try:
        for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            values = [str(value).strip() if value is not None else "" for value in row]
            if all(column in values for column in required_columns):
                return row_number - 1
            if row_number >= max_scan_rows:
                break
    finally:
        workbook.close()

    raise ValueError(
        f"No header containing all required columns was found within the first "
        f"{max_scan_rows} rows."
    )


def read_eprime_excel(file_path: Path) -> pd.DataFrame:
    header_row = find_header_row(file_path, REQUIRED_COLUMNS)
    data = pd.read_excel(file_path, header=header_row, engine="openpyxl")
    data.columns = [str(column).strip().replace("\ufeff", "") for column in data.columns]
    return data


def process_session(
    file_path: Path,
    source_file: str,
    subject: str,
    condition: str,
) -> tuple[
    dict[str, Any] | None,
    list[dict[str, Any]],
    list[dict[str, Any]],
    list[dict[str, Any]],
]:
    try:
        data = read_eprime_excel(file_path)
    except Exception as exc:
        return None, [], [{
            "subject": subject,
            "condition": condition,
            "source_file": source_file,
            "reason": f"Unable to read input file ({type(exc).__name__}).",
        }], []

    missing_columns = [column for column in REQUIRED_COLUMNS if column not in data.columns]
    if missing_columns:
        return None, [], [{
            "subject": subject,
            "condition": condition,
            "source_file": source_file,
            "reason": f"Missing required columns: {missing_columns}",
            "available_columns": ", ".join(data.columns),
        }], []

    working_data = data.copy()
    working_data["_row_in_data"] = working_data.index + 1

    for column in REQUIRED_COLUMNS:
        working_data[column] = pd.to_numeric(working_data[column], errors="coerce")

    retrieval_data = working_data[
        working_data["SearchingPicCRBox"].isin([3, 4])
    ].copy()
    retrieval_trials_original = int(len(retrieval_data))

    retrieval_data["is_no_response"] = retrieval_data["Searching.RESP"].isna()
    retrieval_data["is_fast_response"] = (
        retrieval_data["Searching.RESP"].notna()
        & retrieval_data["Searching.RT"].notna()
        & (retrieval_data["Searching.RT"] < MIN_RT_MS)
    )

    excluded_fast_data = retrieval_data[retrieval_data["is_fast_response"]].copy()
    valid_data = retrieval_data[~retrieval_data["is_fast_response"]].copy()

    valid_data["trial_class"] = valid_data.apply(
        lambda row: classify_trial(row["SearchingPicCRBox"], row["Searching.RESP"]),
        axis=1,
    )
    valid_data["stimulus_label"] = valid_data["SearchingPicCRBox"].map(
        {3: "old", 4: "new"}
    )
    valid_data["response_label"] = valid_data["Searching.RESP"].map(
        {3: "old", 4: "new"}
    )
    valid_data["derived_accuracy"] = valid_data["trial_class"].map({
        "hit": 1,
        "correct_rejection": 1,
        "miss": 0,
        "false_alarm": 0,
        "no_response": pd.NA,
    })

    responded_data = valid_data[valid_data["Searching.RESP"].notna()].copy()

    trial_rows = (
        valid_data.assign(subject=subject, condition=condition)[
            [
                "subject",
                "condition",
                "_row_in_data",
                "stimulus_label",
                "SearchingPicCRBox",
                "response_label",
                "Searching.RESP",
                "Searching.ACC",
                "derived_accuracy",
                "Searching.RT",
                "trial_class",
                "is_no_response",
                "is_fast_response",
            ]
        ]
        .rename(columns={
            "_row_in_data": "row_in_data",
            "SearchingPicCRBox": "stimulus_type",
            "Searching.RESP": "response",
            "Searching.ACC": "eprime_accuracy",
            "Searching.RT": "rt_ms",
        })
        .to_dict("records")
    )

    excluded_fast_rows = (
        excluded_fast_data.assign(subject=subject, condition=condition)[
            [
                "subject",
                "condition",
                "_row_in_data",
                "SearchingPicCRBox",
                "Searching.RESP",
                "Searching.ACC",
                "Searching.RT",
            ]
        ]
        .rename(columns={
            "_row_in_data": "row_in_data",
            "SearchingPicCRBox": "stimulus_type",
            "Searching.RESP": "response",
            "Searching.ACC": "eprime_accuracy",
            "Searching.RT": "rt_ms",
        })
        .assign(exclusion_reason=f"Response RT below {MIN_RT_MS} ms")
        .to_dict("records")
    )

    old_total = int((retrieval_data["SearchingPicCRBox"] == 3).sum())
    new_total = int((retrieval_data["SearchingPicCRBox"] == 4).sum())
    old_total_after_fast_filter = int((valid_data["SearchingPicCRBox"] == 3).sum())
    new_total_after_fast_filter = int((valid_data["SearchingPicCRBox"] == 4).sum())

    hit = int((valid_data["trial_class"] == "hit").sum())
    miss = int((valid_data["trial_class"] == "miss").sum())
    false_alarm = int((valid_data["trial_class"] == "false_alarm").sum())
    correct_rejection = int((valid_data["trial_class"] == "correct_rejection").sum())

    no_response_old = int(
        (
            (valid_data["SearchingPicCRBox"] == 3)
            & (valid_data["trial_class"] == "no_response")
        ).sum()
    )
    no_response_new = int(
        (
            (valid_data["SearchingPicCRBox"] == 4)
            & (valid_data["trial_class"] == "no_response")
        ).sum()
    )
    no_response_total = no_response_old + no_response_new

    old_sdt_total = hit + miss
    new_sdt_total = false_alarm + correct_rejection
    responded_total = old_sdt_total + new_sdt_total

    hit_rate = safe_divide(hit, old_sdt_total)
    false_alarm_rate = safe_divide(false_alarm, new_sdt_total)
    pr = (
        hit_rate - false_alarm_rate
        if pd.notna(hit_rate) and pd.notna(false_alarm_rate)
        else float("nan")
    )

    z_hit = z_score_from_rate(corrected_rate(hit, old_sdt_total))
    z_false_alarm = z_score_from_rate(corrected_rate(false_alarm, new_sdt_total))
    d_prime = (
        z_hit - z_false_alarm
        if pd.notna(z_hit) and pd.notna(z_false_alarm)
        else float("nan")
    )
    bias_c = (
        -0.5 * (z_hit + z_false_alarm)
        if pd.notna(z_hit) and pd.notna(z_false_alarm)
        else float("nan")
    )

    accuracy_mismatch_count = int(
        (
            valid_data["Searching.ACC"].notna()
            & valid_data["derived_accuracy"].notna()
            & (
                valid_data["Searching.ACC"].astype("Int64")
                != valid_data["derived_accuracy"].astype("Int64")
            )
        ).sum()
    )

    reaction_time_metrics: dict[str, Any] = {}
    reaction_time_metrics.update(
        summarize_reaction_time(responded_data["Searching.RT"], "rt_all")
    )
    reaction_time_metrics.update(
        summarize_reaction_time(
            valid_data.loc[valid_data["trial_class"] == "hit", "Searching.RT"],
            "rt_hit",
        )
    )
    reaction_time_metrics.update(
        summarize_reaction_time(
            valid_data.loc[valid_data["trial_class"] == "miss", "Searching.RT"],
            "rt_miss",
        )
    )
    reaction_time_metrics.update(
        summarize_reaction_time(
            valid_data.loc[
                valid_data["trial_class"] == "false_alarm", "Searching.RT"
            ],
            "rt_false_alarm",
        )
    )
    reaction_time_metrics.update(
        summarize_reaction_time(
            valid_data.loc[
                valid_data["trial_class"] == "correct_rejection", "Searching.RT"
            ],
            "rt_correct_rejection",
        )
    )
    reaction_time_metrics.update(
        summarize_reaction_time(
            responded_data.loc[
                responded_data["SearchingPicCRBox"] == 3, "Searching.RT"
            ],
            "rt_old_stimulus",
        )
    )
    reaction_time_metrics.update(
        summarize_reaction_time(
            responded_data.loc[
                responded_data["SearchingPicCRBox"] == 4, "Searching.RT"
            ],
            "rt_new_stimulus",
        )
    )

    summary_row: dict[str, Any] = {
        "subject": subject,
        "condition": condition,
        "source_file": source_file,
        "retrieval_trials_original": retrieval_trials_original,
        "excluded_fast_response_n": int(len(excluded_fast_data)),
        "retrieval_trials_after_fast_filter": int(len(valid_data)),
        "old_total": old_total,
        "new_total": new_total,
        "old_total_after_fast_filter": old_total_after_fast_filter,
        "new_total_after_fast_filter": new_total_after_fast_filter,
        "no_response_old": no_response_old,
        "no_response_new": no_response_new,
        "no_response_total": no_response_total,
        "responded_total": responded_total,
        "old_sdt_total": old_sdt_total,
        "new_sdt_total": new_sdt_total,
        "hit": hit,
        "miss": miss,
        "false_alarm": false_alarm,
        "correct_rejection": correct_rejection,
        "hit_rate": hit_rate,
        "false_alarm_rate": false_alarm_rate,
        "pr": pr,
        "d_prime": d_prime,
        "bias_c": bias_c,
        "accuracy_mismatch_count": accuracy_mismatch_count,
    }
    summary_row.update(reaction_time_metrics)

    return summary_row, trial_rows, [], excluded_fast_rows


def sort_output_tables(
    summary: pd.DataFrame,
    trial_level: pd.DataFrame,
    excluded_fast_rt: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    condition_rank = {condition: rank for rank, condition in enumerate(CONDITION_ORDER)}

    if not summary.empty:
        summary["_condition_rank"] = summary["condition"].map(condition_rank)
        summary = summary.sort_values(["subject", "_condition_rank"]).drop(
            columns="_condition_rank"
        )

    if not trial_level.empty:
        trial_level["_condition_rank"] = trial_level["condition"].map(condition_rank)
        trial_level = trial_level.sort_values(
            ["subject", "_condition_rank", "row_in_data"]
        ).drop(columns="_condition_rank")

    if not excluded_fast_rt.empty:
        excluded_fast_rt["_condition_rank"] = excluded_fast_rt["condition"].map(
            condition_rank
        )
        excluded_fast_rt = excluded_fast_rt.sort_values(
            ["subject", "_condition_rank", "row_in_data"]
        ).drop(columns="_condition_rank")

    return summary, trial_level, excluded_fast_rt


def build_notes_table() -> pd.DataFrame:
    return pd.DataFrame({
        "item": [
            "Header detection",
            "Fast-response exclusion",
            "No response",
            "Hit and false-alarm rates",
            "d-prime",
            "Bias index",
            "Reaction-time summaries",
            "E-Prime accuracy",
        ],
        "value": [
            "The first worksheet is scanned for the required E-Prime column names.",
            f"A trial is excluded only when a response is present and RT is below {MIN_RT_MS} ms.",
            "Missing responses are counted separately and excluded from SDT denominators and RT summaries.",
            "Hit rate = hit/(hit+miss); false-alarm rate = false alarm/(false alarm+correct rejection).",
            "Extreme rates of 0 or 1 are corrected as (count+0.5)/(total+1) before z transformation.",
            "bias_c = 0.5 * [z(hit rate) + z(false-alarm rate)].",
            "RT summaries include responded, non-fast trials only.",
            "Searching.ACC is used only to check agreement with response-derived accuracy.",
        ],
    })


def run(input_directory: Path, output_file: Path) -> None:
    if not input_directory.is_dir():
        raise NotADirectoryError(f"Input directory does not exist: {input_directory}")

    subject_directories = sorted(
        [
            path
            for path in input_directory.iterdir()
            if path.is_dir() and path.name.lower().startswith("sub")
        ],
        key=subject_sort_key,
    )

    if not subject_directories:
        raise FileNotFoundError(
            "No subject directories beginning with 'sub' were found in the input directory."
        )

    summary_rows: list[dict[str, Any]] = []
    trial_rows: list[dict[str, Any]] = []
    excluded_rows: list[dict[str, Any]] = []
    problem_rows: list[dict[str, Any]] = []

    for subject_directory in subject_directories:
        subject = subject_directory.name

        for condition in CONDITION_ORDER:
            file_name = CONDITION_FILES[condition]
            file_path = subject_directory / file_name
            source_file = f"{subject}/{file_name}"

            if not file_path.exists():
                problem_rows.append({
                    "subject": subject,
                    "condition": condition,
                    "source_file": source_file,
                    "reason": "Input file not found.",
                })
                continue

            summary_row, session_trials, problems, session_exclusions = process_session(
                file_path=file_path,
                source_file=source_file,
                subject=subject,
                condition=condition,
            )

            if summary_row is not None:
                summary_rows.append(summary_row)
            trial_rows.extend(session_trials)
            excluded_rows.extend(session_exclusions)
            problem_rows.extend(problems)

    summary = pd.DataFrame(summary_rows)
    trial_level = pd.DataFrame(trial_rows)
    excluded_fast_rt = pd.DataFrame(excluded_rows)
    problems = pd.DataFrame(problem_rows)

    summary, trial_level, excluded_fast_rt = sort_output_tables(
        summary,
        trial_level,
        excluded_fast_rt,
    )

    output_file.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="summary", index=False)
        trial_level.to_excel(writer, sheet_name="trial_level", index=False)
        excluded_fast_rt.to_excel(writer, sheet_name="excluded_fast_rt", index=False)
        problems.to_excel(writer, sheet_name="problems", index=False)
        build_notes_table().to_excel(writer, sheet_name="notes", index=False)

    print(f"Output written to: {output_file}")


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Calculate recognition-memory metrics from E-Prime Excel files."
    )
    parser.add_argument(
        "input_directory",
        type=Path,
        help="Directory containing subject folders such as sub01, sub02, and so on.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("behavior_metrics_all_subjects.xlsx"),
        help="Output Excel file (default: behavior_metrics_all_subjects.xlsx).",
    )
    return parser.parse_args()


def main() -> None:
    arguments = parse_arguments()
    run(arguments.input_directory, arguments.output)


if __name__ == "__main__":
    main()
